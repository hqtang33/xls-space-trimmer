import os

import xlrd
import xlutils.copy
import xlwt
import openpyxl

OUTPUT_PATH = 'output'

def getOutCell(outSheet, colIndex, rowIndex):
        """ HACK: Extract the internal xlwt cell representation. """
        row = outSheet._Worksheet__rows.get(rowIndex)
        if not row: return None
        cell = row._Row__cells.get(colIndex)
        return cell

def setOutCell(outSheet, col, row, value):
    """ Change cell value without changing formatting. """
    previousCell = getOutCell(outSheet, col, row)
    outSheet.write(row, col, value)
    if previousCell:
        newCell = getOutCell(outSheet, col, row)
        if newCell:
            newCell.xf_idx = previousCell.xf_idx
            
def convertValueToStrippedString(value):
    if type(value) == float:
        out = str(int(value))
    elif type(value) == int:
        out = str(value)
    else:
        out = value.strip()
    return out

def processXls(filename):
    print('Processing file: {}'.format(file))
    book = xlrd.open_workbook(file, formatting_info=True)
    new_book = xlutils.copy.copy(book)
    num_of_sheets = book.nsheets
    for i in range(num_of_sheets):
        sheet = book.sheet_by_index(i)
        print('--> Sheet name: {}'.format(sheet.name))
        new_sheet = new_book.get_sheet(i)
        n_col = sheet.ncols
        n_row = sheet.nrows
        for row in range(n_row):
            for col in range(n_col):
                cell_value = convertValueToStrippedString(sheet.cell_value(rowx=row, colx=col))
                if not cell_value == None:
                    cell_value = convertValueToStrippedString(cell_value)
                    setOutCell(new_sheet, col, row, cell_value)
    output_path = os.path.join(OUTPUT_PATH, file)
    new_book.save(output_path)
    print('Saved file: {}'.format(output_path))
    
def processXlsx(file):
    print('Processing file: {}'.format(file))
    workbook = openpyxl.load_workbook(file)
    for index, sheet in enumerate(workbook.worksheets):
        print('--> Sheet name: {}'.format(workbook.sheetnames[index]))
        n_row = sheet.max_row
        n_col = sheet.max_column
        
        for row in range(n_row):
            for col in range(n_col):
                # excel col and row start from 1
                cell = sheet.cell(row+1, col+1)
                cell_value = cell.value
                if not cell_value == None:
                    cell_value = convertValueToStrippedString(cell_value)
                    cell.value = cell_value
    output_path = os.path.join(OUTPUT_PATH, file)
    workbook.save(output_path)
    print('Saved file: {}'.format(output_path))
        
if __name__ == "__main__":
    files = os.listdir()
    files = [file for file in files if file.endswith('.xls') or file.endswith('.xlsx')]
    if not len(files) == 0:
        if not os.path.exists(OUTPUT_PATH):
            os.mkdir(OUTPUT_PATH)
        count = 0
        for file in files:
            if file.endswith('.xls'):
                processXls(file)
            else:
                processXlsx(file)
            print('--------------------------------------')
            count += 1
        print('\n{} files processed...'.format(count))
    else:
        print('No .xls file found!')
    input('Press any key to close...')